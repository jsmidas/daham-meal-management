import openpyxl
from typing import Dict, List, Optional
from pathlib import Path
from decimal import Decimal
from sqlalchemy.orm import Session
from models import Supplier, Ingredient, SupplierIngredient

class PerfectImporter:
    """완벽한 가격 복구 임포터 - 수동 테스트 기반"""
    
    def __init__(self, db: Session):
        self.db = db
        self.supplier_mapping = {
            "82_(사조푸디스트)": "사조푸디스트",
            "82_동원홈푸드": "동원홈푸드", 
            "82_영유통": "영유통",
            "82_현대그린푸드": "현대그린푸드",
            "82다함푸드": "CJ제일제당"
        }
        self.stats = {
            'total_processed': 0,
            'successful_imports': 0,
            'price_restored': 0,
            'errors': 0
        }
    
    def perfect_import_supplier_data(self, file_path: str) -> Dict:
        """검증된 방법으로 완벽 임포트"""
        file_path = Path(file_path)
        
        try:
            supplier_name = self._extract_supplier_name(file_path.name)
            supplier = self._get_or_create_supplier(supplier_name)
            
            print(f"\\n완벽 임포트: {supplier_name}")
            
            # 기존 데이터 삭제
            existing_count = self.db.query(SupplierIngredient).filter(
                SupplierIngredient.supplier_id == supplier.id
            ).count()
            
            if existing_count > 0:
                print(f"  기존 데이터 {existing_count:,}개 삭제...")
                self.db.query(SupplierIngredient).filter(
                    SupplierIngredient.supplier_id == supplier.id
                ).delete()
                self.db.commit()
            
            # Excel 파일 열기
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            print(f"  파일 크기: {ws.max_row-1:,}행")
            
            self.stats = {
                'total_processed': ws.max_row - 1,
                'successful_imports': 0,
                'price_restored': 0,
                'errors': 0
            }
            
            # 데이터 행 처리 (헤더 제외)
            for row_idx in range(2, ws.max_row + 1):
                try:
                    success = self._process_excel_row_perfect(ws, row_idx, supplier.id)
                    if success:
                        self.stats['successful_imports'] += 1
                except Exception as e:
                    self.stats['errors'] += 1
                    if self.stats['errors'] <= 3:  # 처음 3개 오류만 표시
                        print(f"    행 {row_idx} 오류: {str(e)}")
                
                # 진행상황
                if row_idx % 5000 == 0 or row_idx == ws.max_row:
                    processed = row_idx - 1
                    total = ws.max_row - 1
                    print(f"    진행: {processed:,}/{total:,} ({processed/total*100:.1f}%)")
                    self.db.commit()  # 중간 커밋
            
            self.db.commit()
            wb.close()
            
            print(f"  완료: {self.stats['successful_imports']:,}개 임포트")
            print(f"  가격 복구: {self.stats['price_restored']:,}개")
            print(f"  복구율: {self.stats['price_restored']/self.stats['successful_imports']*100:.1f}%")
            
            return {
                'success': True,
                'supplier_name': supplier_name,
                'stats': self.stats
            }
            
        except Exception as e:
            self.db.rollback()
            print(f"  실패: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'stats': self.stats
            }
    
    def _process_excel_row_perfect(self, ws, row_idx: int, supplier_id: int) -> bool:
        """완벽한 행 처리 - 수동 테스트 방법 사용"""
        # 식자재명 (4번째 컬럼)
        ingredient_name = ws.cell(row=row_idx, column=4).value
        if not ingredient_name:
            return False
        
        ingredient_name = str(ingredient_name).strip()
        if not ingredient_name or ingredient_name == 'None':
            return False
        
        # 기본 정보
        ingredient_code = str(ws.cell(row=row_idx, column=3).value or '').strip()
        origin = str(ws.cell(row=row_idx, column=5).value or '').strip()
        unit = str(ws.cell(row=row_idx, column=7).value or 'kg').strip() or 'kg'
        
        # 가격 정보 (10번째, 11번째 컬럼)
        unit_price_raw = ws.cell(row=row_idx, column=10).value
        selling_price_raw = ws.cell(row=row_idx, column=11).value
        
        unit_price = self._parse_price_perfect(unit_price_raw)
        selling_price = self._parse_price_perfect(selling_price_raw)
        
        # 식재료 생성/조회
        ingredient = self._get_or_create_ingredient(ingredient_name, unit, supplier_id)
        
        # 공급업체-식재료 관계 생성
        supplier_ingredient = SupplierIngredient(
            supplier_id=supplier_id,
            ingredient_id=ingredient.id,
            ingredient_code=ingredient_code,
            origin=origin,
            is_published=True,
            unit=unit,
            is_tax_free=False,
            unit_price=unit_price,
            selling_price=selling_price,
            note=''
        )
        
        self.db.add(supplier_ingredient)
        
        # 가격 복구 통계
        if unit_price > 0 or selling_price > 0:
            self.stats['price_restored'] += 1
        
        return True
    
    def _parse_price_perfect(self, value) -> Decimal:
        """검증된 가격 파싱 방법"""
        try:
            if value is None:
                return Decimal('0')
            if isinstance(value, (int, float)):
                return Decimal(str(value))
            
            # 문자열 처리
            value_str = str(value).strip()
            if not value_str or value_str == 'None':
                return Decimal('0')
            
            # 숫자만 추출
            import re
            clean_value = re.sub(r'[^\\d.]', '', value_str)
            return Decimal(clean_value) if clean_value else Decimal('0')
            
        except Exception:
            return Decimal('0')
    
    def _extract_supplier_name(self, filename: str) -> str:
        for key, supplier_name in self.supplier_mapping.items():
            if key in filename:
                return supplier_name
        return filename.split('.')[0]
    
    def _get_or_create_supplier(self, name: str) -> Supplier:
        supplier = self.db.query(Supplier).filter(Supplier.name == name).first()
        if not supplier:
            supplier = Supplier(name=name, update_frequency="2주")
            self.db.add(supplier)
            self.db.commit()
            self.db.refresh(supplier)
        return supplier
    
    def _get_or_create_ingredient(self, name: str, base_unit: str, supplier_id: int) -> Ingredient:
        ingredient = self.db.query(Ingredient).filter(Ingredient.name == name).first()
        if not ingredient:
            ingredient = Ingredient(
                name=name,
                base_unit=base_unit,
                supplier_id=supplier_id
            )
            self.db.add(ingredient)
            self.db.commit()
            self.db.refresh(ingredient)
        return ingredient

def perfect_import_all_files(db: Session):
    """완벽한 모든 파일 임포트"""
    importer = PerfectImporter(db)
    
    print("=== 완벽한 가격 복구 임포트 ===")
    
    upload_dir = Path("sample data/upload")
    results = {}
    total_items = 0
    total_priced = 0
    
    for file_path in upload_dir.glob("*.xls*"):
        if file_path.name != "food_sample.xls":
            result = importer.perfect_import_supplier_data(str(file_path))
            results[file_path.name] = result
            
            if result['success']:
                total_items += result['stats']['successful_imports']
                total_priced += result['stats']['price_restored']
    
    print(f"\\n=== 최종 완벽 결과 ===")
    print(f"총 임포트: {total_items:,}개")
    print(f"가격 복구: {total_priced:,}개")
    print(f"가격 복구율: {total_priced/total_items*100:.1f}%" if total_items > 0 else "N/A")
    
    return results

if __name__ == "__main__":
    print("완벽한 임포터 준비 완료")
    print("사용법: from perfect_importer import perfect_import_all_files; perfect_import_all_files(db)")