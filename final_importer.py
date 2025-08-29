import openpyxl
import pandas as pd
from typing import Dict, List, Tuple, Optional
from pathlib import Path
from decimal import Decimal
from sqlalchemy.orm import Session
from models import Supplier, Ingredient, SupplierIngredient

class FinalImporter:
    """openpyxl을 사용한 최종 수정 임포터"""
    
    def __init__(self, db: Session):
        self.db = db
        self.supplier_mapping = {
            "82_(사조푸디스트)": "사조푸디스트",
            "82_동원홈푸드": "동원홈푸드", 
            "82_영유통": "영유통",
            "82_현대그린푸드": "현대그린푸드",
            "82다함푸드": "CJ제일제당"
        }
        self.stats = {
            'total_processed': 0,
            'successful_imports': 0,
            'price_restored': 0,
            'errors': 0
        }
    
    def final_import_supplier_data(self, file_path: str) -> Dict:
        """openpyxl로 가격 정보까지 완벽 임포트"""
        file_path = Path(file_path)
        
        try:
            supplier_name = self._extract_supplier_name(file_path.name)
            supplier = self._get_or_create_supplier(supplier_name)
            
            print(f"처리 중: {supplier_name}")
            
            # 기존 데이터 삭제
            existing_count = self.db.query(SupplierIngredient).filter(
                SupplierIngredient.supplier_id == supplier.id
            ).count()
            
            if existing_count > 0:
                print(f"  기존 데이터 {existing_count:,}개 삭제...")
                self.db.query(SupplierIngredient).filter(
                    SupplierIngredient.supplier_id == supplier.id
                ).delete()
                self.db.commit()
            
            # openpyxl로 파일 읽기
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # 컬럼 매핑 (첫 번째 행에서 헤더 확인)
            headers = []
            for col_idx in range(1, ws.max_column + 1):
                header_value = ws.cell(row=1, column=col_idx).value
                headers.append(header_value)
            
            print(f"  컬럼 구조: {headers}")
            
            # 컬럼 인덱스 찾기
            column_indices = self._find_column_indices(headers)
            
            self.stats = {
                'total_processed': ws.max_row - 1,  # 헤더 제외
                'successful_imports': 0,
                'price_restored': 0,
                'errors': 0
            }
            
            # 데이터 행 처리 (2행부터 시작, 1행은 헤더)
            for row_idx in range(2, ws.max_row + 1):
                try:
                    self._process_excel_row(ws, row_idx, column_indices, supplier.id)
                except Exception as e:
                    self.stats['errors'] += 1
                    if self.stats['errors'] <= 5:  # 처음 5개 오류만 출력
                        print(f"    행 {row_idx} 오류: {str(e)}")
                
                # 진행상황 출력
                if row_idx % 5000 == 0 or row_idx == ws.max_row:
                    processed = row_idx - 1
                    total = ws.max_row - 1
                    print(f"    진행: {processed:,}/{total:,} ({processed/total*100:.1f}%)")
                    self.db.commit()
            
            self.db.commit()
            wb.close()
            
            print(f"  완료: {self.stats['successful_imports']:,}개 임포트")
            print(f"  가격 복구: {self.stats['price_restored']:,}개")
            print(f"  오류: {self.stats['errors']:,}개")
            
            return {
                'success': True,
                'supplier_name': supplier_name,
                'stats': self.stats
            }
            
        except Exception as e:
            self.db.rollback()
            return {
                'success': False,
                'error': str(e),
                'stats': self.stats
            }
    
    def _find_column_indices(self, headers: List) -> Dict:
        """헤더에서 컬럼 인덱스 찾기"""
        indices = {}
        
        for i, header in enumerate(headers):
            if header == '식자재명':
                indices['ingredient_name'] = i
            elif header == '식자재코드':
                indices['ingredient_code'] = i
            elif header == '단위':
                indices['unit'] = i
            elif header == '입고단가':
                indices['unit_price'] = i
            elif header == '판매단가':
                indices['selling_price'] = i
            elif header == '원산지':
                indices['origin'] = i
            elif header == '게시여부':
                indices['is_published'] = i
            elif header == '면세여부':
                indices['is_tax_free'] = i
            elif header == '비고':
                indices['note'] = i
        
        return indices
    
    def _process_excel_row(self, ws, row_idx: int, column_indices: Dict, supplier_id: int):
        """Excel 행을 직접 처리"""
        # 식재료명 확인
        if 'ingredient_name' not in column_indices:
            raise Exception("식자재명 컬럼을 찾을 수 없음")
        
        ingredient_name_cell = ws.cell(row=row_idx, column=column_indices['ingredient_name'] + 1)
        ingredient_name = str(ingredient_name_cell.value).strip() if ingredient_name_cell.value else ""
        
        if not ingredient_name or ingredient_name == 'None':
            return  # 빈 행 스킵
        
        # 다른 컬럼 값 추출
        def get_cell_value(col_key: str, default=""):
            if col_key in column_indices:
                cell = ws.cell(row=row_idx, column=column_indices[col_key] + 1)
                return cell.value if cell.value is not None else default
            return default
        
        unit = str(get_cell_value('unit', 'kg')).strip() or 'kg'
        
        # 가격 정보 - 직접 셀에서 가져오기
        unit_price_raw = get_cell_value('unit_price', 0)
        selling_price_raw = get_cell_value('selling_price', 0)
        
        unit_price = self._parse_number(unit_price_raw)
        selling_price = self._parse_number(selling_price_raw)
        
        # 식재료 생성/조회
        ingredient = self._get_or_create_ingredient(ingredient_name, unit, supplier_id)
        
        # 공급업체-식재료 관계 생성
        supplier_ingredient = SupplierIngredient(
            supplier_id=supplier_id,
            ingredient_id=ingredient.id,
            ingredient_code=str(get_cell_value('ingredient_code', '')).strip(),
            origin=str(get_cell_value('origin', '')).strip(),
            is_published=self._parse_boolean(get_cell_value('is_published', True)),
            unit=unit,
            is_tax_free=self._parse_boolean(get_cell_value('is_tax_free', False)),
            unit_price=unit_price,
            selling_price=selling_price,
            note=str(get_cell_value('note', '')).strip()
        )
        
        self.db.add(supplier_ingredient)
        self.stats['successful_imports'] += 1
        
        if unit_price > 0 or selling_price > 0:
            self.stats['price_restored'] += 1
    
    def _parse_number(self, value) -> Decimal:
        """숫자값을 Decimal로 파싱"""
        try:
            if value is None:
                return Decimal('0')
            if isinstance(value, (int, float)):
                return Decimal(str(value))
            
            # 문자열인 경우
            value_str = str(value).strip()
            if not value_str or value_str == 'None':
                return Decimal('0')
            
            # 숫자가 아닌 문자 제거
            import re
            clean_value = re.sub(r'[^\d.]', '', value_str)
            
            return Decimal(clean_value) if clean_value else Decimal('0')
            
        except Exception:
            return Decimal('0')
    
    def _parse_boolean(self, value) -> bool:
        """불린값 파싱"""
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ['true', '1', 'y', 'yes', '예', '참']
        return bool(value)
    
    def _extract_supplier_name(self, filename: str) -> str:
        """파일명에서 공급업체 이름 추출"""
        for key, supplier_name in self.supplier_mapping.items():
            if key in filename:
                return supplier_name
        return filename.split('.')[0]
    
    def _get_or_create_supplier(self, name: str) -> Supplier:
        """공급업체 생성/조회"""
        supplier = self.db.query(Supplier).filter(Supplier.name == name).first()
        if not supplier:
            supplier = Supplier(name=name, update_frequency="2주")
            self.db.add(supplier)
            self.db.commit()
            self.db.refresh(supplier)
        return supplier
    
    def _get_or_create_ingredient(self, name: str, base_unit: str, supplier_id: int) -> Ingredient:
        """식재료 생성/조회"""
        ingredient = self.db.query(Ingredient).filter(Ingredient.name == name).first()
        if not ingredient:
            ingredient = Ingredient(
                name=name,
                base_unit=base_unit,
                supplier_id=supplier_id
            )
            self.db.add(ingredient)
            self.db.commit()
            self.db.refresh(ingredient)
        return ingredient

def final_import_all_files(db: Session):
    """최종 완벽 임포트 실행"""
    importer = FinalImporter(db)
    
    print("=== 최종 완벽 임포트 (가격 정보 복구) ===")
    print()
    
    upload_dir = Path("sample data/upload")
    results = {}
    total_items = 0
    total_priced = 0
    
    for file_path in upload_dir.glob("*.xls*"):
        if file_path.name != "food_sample.xls":
            result = importer.final_import_supplier_data(str(file_path))
            results[file_path.name] = result
            
            if result['success']:
                total_items += result['stats']['successful_imports']
                total_priced += result['stats']['price_restored']
    
    print(f"\n=== 최종 결과 ===")
    print(f"총 임포트: {total_items:,}개")
    print(f"가격 정보 복구: {total_priced:,}개")
    print(f"가격 복구율: {total_priced/total_items*100:.1f}%" if total_items > 0 else "N/A")
    
    return results

if __name__ == "__main__":
    print("최종 임포터 준비 완료")
    print("사용법: from final_importer import final_import_all_files; final_import_all_files(db)")